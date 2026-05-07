from __future__ import annotations

import io
import re

from django.conf import settings
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from apps.dashboard.filters import FeedbackFilterSet
from apps.dashboard.services.analytics_engine import AnalyticsEngine
from apps.feedback.models import Feedback


class ReportGenerator:
    @staticmethod
    def content_type(format: str) -> str:
        if format == "pdf":
            return "application/pdf"
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def generate(self, format: str, filters: dict, template_id: str, user):
        queryset = self._queryset(filters)
        filename = f"refuconnect_{template_id}_{timezone.localdate()}.{format}"
        if format == "pdf":
            return self._generate_pdf(queryset, filters, template_id, user), filename
        return self._generate_excel(queryset, filters), filename

    def row_count(self, filters: dict) -> int:
        return self._queryset(filters).count()

    def _queryset(self, filters: dict):
        qs = Feedback.objects.select_related("sentiment").prefetch_related(
            "feedback_categories__category"
        )
        filterset = FeedbackFilterSet(data=filters, queryset=qs)
        return filterset.qs.distinct() if filterset.is_valid() else qs

    def _generate_pdf(self, queryset, filters: dict, template_id: str, user) -> bytes:
        from weasyprint import HTML

        samples = [
            {
                "feedback_id": fb.feedback_id,
                "submitted_at": fb.submitted_at,
                "channel": fb.channel,
                "message": self._anonymise(fb.message_text_en or fb.message_text or ""),
            }
            for fb in queryset.order_by("-submitted_at")[:5]
        ]
        analytics = AnalyticsEngine().get_summary(filters, getattr(user, "organisation", "") or 1)
        context = {
            "organisation_name": getattr(user, "organisation", "") or "RefuConnect",
            "logo_path": getattr(settings, "REPORT_LOGO_PATH", ""),
            "title": "RefuConnect Feedback Report",
            "generated_at": timezone.now(),
            "filters": filters,
            "analytics": analytics,
            "samples": samples,
            "sentiment_trend_svg": self._build_sentiment_trend_svg(
                analytics.get("sentiment_trend", [])
            ),
        }
        template = f"reports/{template_id}.html"
        html_string = render_to_string(template, context)
        return HTML(string=html_string, base_url=str(settings.BASE_DIR)).write_pdf()

    def _build_sentiment_trend_svg(self, rows: list[dict]) -> str:
        if not rows:
            return ""

        chart_rows = rows[-14:]
        width = 760
        height = 260
        margin_left = 45
        margin_right = 16
        margin_top = 16
        margin_bottom = 48
        plot_width = width - margin_left - margin_right
        plot_height = height - margin_top - margin_bottom
        bar_count = max(len(chart_rows), 1)
        slot_width = plot_width / bar_count
        bar_width = max(slot_width * 0.62, 8)

        totals = [
            row.get("Positive", 0)
            + row.get("Neutral", 0)
            + row.get("Negative", 0)
            + row.get("Uncertain", 0)
            for row in chart_rows
        ]
        y_max = max(max(totals), 1)

        palette = {
            "Positive": "#2A9D8F",
            "Neutral": "#E9C46A",
            "Negative": "#E76F51",
            "Uncertain": "#7B8FA1",
        }

        parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
            '<rect width="100%" height="100%" fill="white"/>',
            f'<line x1="{margin_left}" y1="{margin_top + plot_height}" x2="{margin_left + plot_width}" y2="{margin_top + plot_height}" stroke="#B9C5D1" stroke-width="1"/>',
        ]

        for index, row in enumerate(chart_rows):
            x = margin_left + (index * slot_width) + (slot_width - bar_width) / 2
            y_cursor = margin_top + plot_height
            for sentiment in ("Positive", "Neutral", "Negative", "Uncertain"):
                value = int(row.get(sentiment, 0) or 0)
                if value <= 0:
                    continue
                rect_height = (value / y_max) * plot_height
                y_cursor -= rect_height
                parts.append(
                    f'<rect x="{x:.2f}" y="{y_cursor:.2f}" width="{bar_width:.2f}" height="{rect_height:.2f}" fill="{palette[sentiment]}"/>'
                )

            label = str(row.get("date", ""))[5:10]
            label_x = x + bar_width / 2
            parts.append(
                f'<text x="{label_x:.2f}" y="{height - 20}" font-size="10" text-anchor="middle" fill="#4A5A6A">{label}</text>'
            )

        legend_x = margin_left
        legend_y = height - 8
        legend_offset = 0
        for sentiment in ("Positive", "Neutral", "Negative", "Uncertain"):
            color = palette[sentiment]
            parts.append(
                f'<rect x="{legend_x + legend_offset}" y="{legend_y - 10}" width="10" height="10" fill="{color}"/>'
            )
            parts.append(
                f'<text x="{legend_x + legend_offset + 14}" y="{legend_y}" font-size="10" fill="#4A5A6A">{sentiment}</text>'
            )
            legend_offset += 95

        parts.append("</svg>")
        return "".join(parts)

    def _generate_excel(self, queryset, filters: dict) -> bytes:
        analytics = AnalyticsEngine().get_summary(filters, 1)
        wb = Workbook()
        summary = wb.active
        assert summary is not None
        summary.title = "Summary"
        self._write_summary(summary, analytics)

        raw = wb.create_sheet("Raw Data")
        raw_headers = [
            "Reference ID",
            "Submitted At (UTC)",
            "Channel",
            "Language",
            "Sentiment",
            "Urgency",
            "Categories",
            "Location",
            "Status",
            "Original Message",
            "English Translation",
        ]
        raw.append(raw_headers)
        for fb in queryset.order_by("-submitted_at"):
            raw.append(
                [
                    fb.feedback_id,
                    fb.submitted_at.strftime("%Y-%m-%d %H:%M:%S UTC"),
                    fb.channel,
                    fb.language,
                    fb.sentiment.sentiment_label if fb.sentiment else "",
                    fb.urgency_level,
                    ", ".join(
                        fc.category.category_name for fc in fb.feedback_categories.all()
                    ),
                    fb.location or "",
                    fb.status,
                    fb.message_text or "",
                    fb.message_text_en or "",
                ]
            )

        trends = wb.create_sheet("Trends")
        trends.append(["Date", "Positive", "Neutral", "Negative", "Uncertain"])
        for row in analytics["sentiment_trend"]:
            trends.append(
                [
                    row["date"],
                    row["Positive"],
                    row["Neutral"],
                    row["Negative"],
                    row["Uncertain"],
                ]
            )

        categories = wb.create_sheet("Categories")
        categories.append(["Category", "Count", "Percentage"])
        for row in analytics["top_categories"]:
            categories.append([row["category_name"], row["count"], row["percentage"]])

        for ws in wb.worksheets:
            self._format_sheet(ws)
        self._add_table(raw, "RawDataTable")
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _write_summary(self, ws, analytics: dict) -> None:
        rows = [
            ("Metric", "Value"),
            ("Total feedback", analytics["volume"]["total"]),
            ("Today", analytics["volume"]["today"]),
            ("This week", analytics["volume"]["this_week"]),
            ("This month", analytics["volume"]["this_month"]),
            ("Urgent open alerts", analytics["urgent_open_count"]),
            ("Unprocessed", analytics["unprocessed_count"]),
        ]
        for row in rows:
            ws.append(row)
        ws.append([])
        ws.append(["Sentiment", "Count", "Percentage"])
        for label, values in analytics["sentiment_distribution"].items():
            ws.append([label, values["count"], values["percentage"]])

    def _format_sheet(self, ws) -> None:
        header_fill = PatternFill("solid", fgColor="264653")
        header_font = Font(bold=True, color="FFFFFF")
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_length + 2, 15), 50)

    def _add_table(self, ws, name: str) -> None:
        if ws.max_row < 2:
            return
        ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    def _anonymise(self, text: str) -> str:
        return re.sub(r"\+?\d[\d\s-]{6,}\d", "[redacted]", text)


def generate_pdf_report(queryset, title: str) -> bytes:
    return ReportGenerator()._generate_pdf(queryset, {}, "executive_summary", None)


def generate_excel_report(queryset) -> bytes:
    return ReportGenerator()._generate_excel(queryset, {})
